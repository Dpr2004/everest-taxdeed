"""Spreadsheet Writer - le lotes do DB e atualiza a planilha dos 11 condados.

Mapeia os campos do DB para as colunas da aba de cada condado, mantendo
as formulas e estilos da planilha intactos.
"""
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from src.db.connection import cursor
from src.workers.base import BaseWorker

PLANILHA_PATH = os.environ.get(
    "SPREADSHEET_PATH",
    "./data/Planilha_Padrao_11Condados_Everest.xlsx",
)

# Mapeamento: DB column -> Excel column letter na aba do condado
# Baseado na estrutura da aba LEVY original
MAP = {
    "parcel_id": "B",
    "case_num": "C",
    "legal_description": "H",
    "address": "L",
    "min_bid": "P",
    "just_value": "Q",
    "assessed_value": "R",
}

PRIMEIRA_LINHA = 4
ULTIMA_LINHA = 40


class SpreadsheetWriter(BaseWorker):
    name = "spreadsheet_writer"

    def __init__(self, county_code=None):
        super().__init__()
        self.county_code = county_code

    def execute(self):
        if not Path(PLANILHA_PATH).exists():
            raise FileNotFoundError(
                f"Planilha nao encontrada em {PLANILHA_PATH}. "
                f"Copie a planilha para este path antes de rodar."
            )

        with cursor() as cur:
            if self.county_code:
                cur.execute(
                    "SELECT * FROM counties WHERE codigo = ? AND ativo = 1",
                    (self.county_code,),
                )
            else:
                cur.execute("SELECT * FROM counties WHERE ativo = 1")
            condados = cur.fetchall()

        wb = load_workbook(PLANILHA_PATH)
        for c in condados:
            aba = c["aba_planilha"]
            if aba not in wb.sheetnames:
                self.logger.warning(f"Aba '{aba}' nao existe na planilha - pulando")
                continue

            # Pegar lotes do proximo sale deste condado
            with cursor() as cur2:
                cur2.execute("""
                    SELECT l.* FROM lots l
                    JOIN sales s ON s.id = l.sale_id
                    WHERE s.county_id = ? AND s.sale_date >= DATE('now')
                    ORDER BY s.sale_date ASC, l.min_bid DESC
                    LIMIT ?
                """, (c["id"], ULTIMA_LINHA - PRIMEIRA_LINHA + 1))
                lotes = cur2.fetchall()

            if not lotes:
                self.logger.info(f"{c['codigo']}: sem lotes no DB - pulando aba")
                continue

            ws = wb[aba]
            # Atualizar titulo A1 com data do sale
            # (opcional, mantem original se preferir)
            for idx, lot in enumerate(lotes):
                linha = PRIMEIRA_LINHA + idx
                for field, col in MAP.items():
                    valor = lot[field] if field in lot.keys() else None
                    if valor is not None:
                        cell = ws[f"{col}{linha}"]
                        cell.value = valor
                        # Marcar inputs em azul (padrao do modelo)
                        cell.font = Font(name="Arial", size=10, color="0000FF")
                self.items_processed += 1
            self.logger.info(f"{c['codigo']}: {len(lotes)} linhas atualizadas na aba")

        wb.save(PLANILHA_PATH)
        self.logger.info(f"Planilha salva: {PLANILHA_PATH}")


if __name__ == "__main__":
    import sys
    code = sys.argv[1] if len(sys.argv) > 1 else None
    SpreadsheetWriter(county_code=code).run()
